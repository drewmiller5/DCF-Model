"""
Stock Data Updater for DCF_Model.xlsx
Usage:  python update_dcf.py AAPL
        python update_dcf.py META --output path/to/DCF_Model.xlsx
Pulls live data from Yahoo Finance and populates the Stock Search sheet.
Requires: pip install yfinance openpyxl
"""

import sys
import os
import argparse
import datetime

try:
    import yfinance as yf
except ImportError:
    print("yfinance not installed. Installing...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "yfinance"])
    import yfinance as yf

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

INPUT_GOLD = "FFF9C4"

def fill(hex_color):
    from openpyxl.styles import PatternFill
    return PatternFill("solid", fgColor=hex_color)

def fetch_stock_data(ticker: str) -> dict:
    """Pull key metrics from Yahoo Finance."""
    tk  = yf.Ticker(ticker)
    try:
        inf = tk.info
    except Exception:
        inf = {}

    if not inf or inf.get("quoteType") is None:
        print(f"\n  ERROR: '{ticker}' not found on Yahoo Finance.")
        print("  Check the ticker symbol and try again (e.g. AAPL, MSFT, NVDA).")
        sys.exit(1)

    def safe(key, default=None):
        val = inf.get(key, default)
        return val if val is not None else default

    # Historical price range
    hist = tk.history(period="1y")
    hi52 = float(hist["High"].max()) if not hist.empty else safe("fiftyTwoWeekHigh", 0)
    lo52 = float(hist["Low"].min())  if not hist.empty else safe("fiftyTwoWeekLow",  0)

    shares = safe("sharesOutstanding", 0)
    price  = safe("currentPrice") or safe("regularMarketPrice") or safe("previousClose", 0)
    mktcap = safe("marketCap", 0)
    ebitda = safe("ebitda", 0)
    rev    = safe("totalRevenue", 0)
    ni     = safe("netIncomeToCommon", 0)
    debt   = safe("totalDebt", 0)
    cash   = safe("totalCash", 0)
    pe     = safe("trailingPE", 0)
    ev_ebitda = safe("enterpriseToEbitda", 0)
    ps     = safe("priceToSalesTrailing12Months", 0)
    div_yield = safe("dividendYield", 0)
    rev_growth = safe("revenueGrowth", 0)
    eps_growth = safe("earningsGrowth", 0)
    beta   = safe("beta", 1.0)
    float_pct = safe("floatShares", shares) / shares if shares else 0.99

    return {
        "ticker":      ticker.upper(),
        "name":        safe("longName", safe("shortName", ticker)),
        "exchange":    safe("exchange", ""),
        "sector":      safe("sector", ""),
        "industry":    safe("industryDisp", safe("industry", "")),
        "price":       price,
        "hi52":        hi52,
        "lo52":        lo52,
        "mktcap_m":    mktcap / 1e6 if mktcap else 0,
        "shares_m":    shares / 1e6 if shares else 0,
        "float_pct":   float_pct,
        "pe":          pe,
        "ev_ebitda":   ev_ebitda,
        "ps":          ps,
        "revenue_m":   rev / 1e6 if rev else 0,
        "ebitda_m":    ebitda / 1e6 if ebitda else 0,
        "ni_m":        ni / 1e6 if ni else 0,
        "debt_m":      debt / 1e6 if debt else 0,
        "cash_m":      cash / 1e6 if cash else 0,
        "net_debt_m":  (debt - cash) / 1e6 if debt and cash else 0,
        "div_yield":   div_yield or 0,
        "rev_growth":  rev_growth or 0,
        "eps_growth":  eps_growth or 0,
        "beta":        beta,
        "updated":     datetime.date.today().strftime("%Y-%m-%d"),
    }

def update_excel(path: str, data: dict):
    """Write fetched data into the DCF Excel model."""
    wb = load_workbook(path)

    if "Stock Search" not in wb.sheetnames:
        print(f"ERROR: 'Stock Search' sheet not found in {path}")
        return

    ws = wb["Stock Search"]
    gold = fill(INPUT_GOLD)

    def write(cell_ref, value):
        c = ws[cell_ref]
        c.value = value
        c.fill  = gold

    # Stock Search sheet layout (5 columns B-E):
    #   Row 4: headers; Row 5: ticker/price/name/updated
    #   Row 7: PRICE & RANGE hdr;  Row 8: price/hi52/lo52
    #   Row 9: SHARES hdr;         Row 10: mktcap/shares/float
    #   Row 11: INCOME hdr;        Row 12: revenue/ebitda/net_income
    #   Row 13: BALANCE SHEET hdr; Row 14: debt/cash/(net_debt formula)
    #   Row 15: VALUATION hdr;     Row 16: pe/ev_ebitda/beta

    write("B5", data["ticker"])
    write("C5", round(data["price"], 2))
    write("D5", data["name"])
    write("E5", data["updated"])

    write("C8",  round(data["price"],    2))
    write("D8",  round(data["hi52"],     2))
    write("E8",  round(data["lo52"],     2))

    write("C10", round(data["mktcap_m"],  1))
    write("D10", round(data["shares_m"],  1))
    write("E10", round(data["float_pct"], 4))

    write("C12", round(data["revenue_m"], 0))
    write("D12", round(data["ebitda_m"],  0))
    write("E12", round(data["ni_m"],      0))

    write("C14", round(data["debt_m"],    0))
    write("D14", round(data["cash_m"],    0))
    # E14 is =C14-D14 formula — leave it

    write("C16", round(data["pe"],        1) if data["pe"] else "N/A")
    write("D16", round(data["ev_ebitda"], 1) if data["ev_ebitda"] else "N/A")
    write("E16", round(data["beta"],      2) if data["beta"] else 1.0)

    # Assumptions sheet layout:
    #   Row 5: name(C5) / ticker(E5)
    #   Row 7: price(C7) / shares(E7)
    #   Row 8: debt(C8) / cash(E8)
    #   Row 22: beta(C22)
    #   Row 28: base revenue(C28)
    if "Assumptions" in wb.sheetnames:
        wa = wb["Assumptions"]
        wa["C5"].value  = data["name"]
        wa["E5"].value  = data["ticker"]
        wa["C7"].value  = round(data["price"],    2)
        wa["E7"].value  = round(data["shares_m"], 1)
        wa["C8"].value  = round(data["debt_m"],   0)
        wa["E8"].value  = round(data["cash_m"],   0)
        wa["C22"].value = round(data["beta"], 2) if data["beta"] else 1.0
        wa["C28"].value = round(data["revenue_m"], 0)

    try:
        wb.save(path)
    except PermissionError:
        print(f"\n  ERROR: Could not save — {os.path.basename(path)} is open in Excel.")
        print("  Close the file and run the command again.")
        sys.exit(1)
    print(f"\nUpdated {path}")
    print(f"  Ticker:  {data['ticker']} -- {data['name']}")
    print(f"  Price:   ${data['price']:.2f}  |  Mkt Cap: ${data['mktcap_m']:,.0f}M")
    print(f"  Revenue: ${data['revenue_m']:,.0f}M  |  EBITDA: ${data['ebitda_m']:,.0f}M")
    print(f"  Net Debt:${data['net_debt_m']:,.0f}M  |  Beta: {data['beta']:.2f}")
    print(f"\n  Open DCF_Model.xlsx and fill in the Assumptions tab gold cells.")
    print(f"  Adjust growth rates and margins to reflect your thesis.")


def main():
    parser = argparse.ArgumentParser(
        description="Update DCF_Model.xlsx with live stock data from Yahoo Finance"
    )
    parser.add_argument("ticker", nargs="?", default=None,
                        help="Stock ticker symbol (e.g. AAPL, MSFT, NVDA)")
    parser.add_argument(
        "--output", "-o",
        default=None,
        help="Path to DCF_Model.xlsx (default: auto-detect in current directory)"
    )
    args = parser.parse_args()

    if not args.ticker:
        args.ticker = input("Enter ticker symbol (e.g. AAPL, MSFT, NVDA): ").strip().upper()
        if not args.ticker:
            print("No ticker provided. Exiting.")
            sys.exit(1)

    if args.output is None:
        import glob
        all_files = sorted(glob.glob(os.path.join(os.getcwd(), "DCF_Model*.xlsx")),
                           key=os.path.getmtime, reverse=True)
        if not all_files:
            print("ERROR: No DCF_Model*.xlsx found in current directory.")
            print("Run  python build_dcf.py  first to create the model.")
            sys.exit(1)

        ticker_up = args.ticker.upper()
        # Prefer a file that contains the ticker in its name
        ticker_matches = [f for f in all_files if f"_{ticker_up}_" in os.path.basename(f)
                          or os.path.basename(f) == f"DCF_Model_{ticker_up}.xlsx"]
        plain = [f for f in all_files if os.path.basename(f) == "DCF_Model.xlsx"]

        if ticker_matches:
            args.output = ticker_matches[0]
        elif plain:
            args.output = plain[0]
        else:
            # Only files exist with other tickers in the name — warn before overwriting
            print(f"\n  WARNING: No file found for {ticker_up}.")
            print(f"  Most recent file: {os.path.basename(all_files[0])}")
            ans = input("  Update it anyway? (y/N): ").strip().lower()
            if ans != "y":
                print("  Aborted. Run  python build_dcf.py  to build a fresh model.")
                sys.exit(0)
            args.output = all_files[0]

    print(f"Fetching data for {args.ticker.upper()} from Yahoo Finance...")
    data = fetch_stock_data(args.ticker)
    update_excel(args.output, data)


if __name__ == "__main__":
    main()
