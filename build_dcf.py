"""
Professional DCF Model Builder - v2 (correct cross-sheet references via cell tracking)
Generates a comprehensive Excel DCF valuation model with stock data integration.
Usage: python build_dcf.py [TICKER]
Output: ~/Desktop/DCF_Model.xlsx (or ~/projects/DCF_Model.xlsx on Windows)
"""

import os, sys, datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

# ─── COLORS ──────────────────────────────────────────────────────────────────
NAVY       = "1B2A4A"
DARK_BLUE  = "2E4170"
MED_BLUE   = "3F5FA0"
LIGHT_BLUE = "D6E4F7"
INPUT_GOLD = "FFF9C4"
CALC_WHITE = "FFFFFF"
GRAY_LIGHT = "F2F2F2"
GREEN_POS  = "C6EFCE"
RED_NEG    = "FFC7CE"
BORDER_CLR = "8EA9C1"
GOLD_ACC   = "C9A22C"

# ─── STYLES ───────────────────────────────────────────────────────────────────
def F(bold=False, size=10, color="000000", italic=False):
    return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)

def Fill(hex_c):
    return PatternFill("solid", fgColor=hex_c)

def B(l=None, r=None, t=None, b=None):
    s = Side(style="thin", color=BORDER_CLR)
    m = Side(style="medium", color=NAVY)
    n = Side(style=None)
    def x(v): return s if v == "t" else m if v == "m" else n
    return Border(left=x(l), right=x(r), top=x(t), bottom=x(b))

thin_border = Border(
    left=Side(style="thin", color=BORDER_CLR),
    right=Side(style="thin", color=BORDER_CLR),
    top=Side(style="thin", color=BORDER_CLR),
    bottom=Side(style="thin", color=BORDER_CLR),
)
thick_l = Border(
    left=Side(style="medium", color=NAVY),
    right=Side(style="thin", color=BORDER_CLR),
    top=Side(style="thin", color=BORDER_CLR),
    bottom=Side(style="thin", color=BORDER_CLR),
)
L = Alignment(horizontal="left",   vertical="center")
R = Alignment(horizontal="right",  vertical="center")
C = Alignment(horizontal="center", vertical="center")

FMT_USD  = '#,##0.0'
FMT_USD2 = '#,##0.00'
FMT_PCT  = '0.0%'
FMT_PCT2 = '0.00%'
FMT_MULT = '0.0'
FMT_INT  = '#,##0'
FMT_SHR  = '#,##0.0'

def sc(ws, ref, val=None, fmt=None, bold=False, size=10, fc="000000",
       bg=None, al=None, italic=False, bdr=None):
    """Style a cell."""
    c = ws[ref] if isinstance(ref, str) else ref
    if val is not None: c.value = val
    c.font = F(bold=bold, size=size, color=fc, italic=italic)
    if bg:  c.fill = Fill(bg)
    if fmt: c.number_format = fmt
    if al:  c.alignment = al
    if bdr: c.border = bdr
    return c

def hdr_cell(ws, ref, val, bg=NAVY, size=10):
    return sc(ws, ref, val, bold=True, size=size, fc="FFFFFF", bg=bg, al=C, bdr=thin_border)

def inp(ws, ref, val=None, fmt=None):
    return sc(ws, ref, val, fmt=fmt, bg=INPUT_GOLD, al=R, bdr=thin_border)

def cal(ws, ref, formula=None, fmt=None, bold=False):
    return sc(ws, ref, formula, fmt=fmt, bold=bold, bg=CALC_WHITE, al=R, bdr=thin_border)

def lbl(ws, ref, val, bold=False, bg=GRAY_LIGHT, italic=False):
    return sc(ws, ref, val, bold=bold, size=9, bg=bg, al=L, bdr=thin_border, italic=italic)

def sec(ws, row, c1, c2, label, bg=DARK_BLUE):
    ref1 = f"{get_column_letter(c1)}{row}"
    ref2 = f"{get_column_letter(c2)}{row}"
    ws.merge_cells(f"{ref1}:{ref2}")
    sc(ws, ref1, f"  {label}", bold=True, size=9, fc="FFFFFF", bg=bg, al=L, bdr=thin_border)

def row_h(ws, row, h=17):
    ws.row_dimensions[row].height = h


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN BUILDER
# ═══════════════════════════════════════════════════════════════════════════════
def build_dcf(filename="DCF_Model.xlsx"):
    wb = Workbook()
    CELLS = {}  # stores key cell addresses for cross-sheet references

    ws_guide  = wb.active; ws_guide.title = "Guide"
    ws_stock  = wb.create_sheet("Stock Search")
    ws_assum  = wb.create_sheet("Assumptions")
    ws_is     = wb.create_sheet("Income Statement")
    ws_wacc   = wb.create_sheet("WACC")
    ws_dcf    = wb.create_sheet("DCF Valuation")
    ws_sens   = wb.create_sheet("Sensitivity")

    for ws, color in [
        (ws_guide, "808080"), (ws_stock, "3F5FA0"), (ws_assum, GOLD_ACC),
        (ws_is, "2E4170"), (ws_wacc, "006100"), (ws_dcf, NAVY), (ws_sens, "9C0006"),
    ]:
        ws.sheet_properties.tabColor = color
        ws.sheet_view.showGridLines = False

    build_guide(ws_guide)
    build_stock(ws_stock, CELLS)
    build_assum(ws_assum, CELLS)   # populates CELLS with Assumptions addresses
    build_is(ws_is, CELLS)         # uses CELLS, populates IS row addresses
    build_wacc(ws_wacc, CELLS)     # uses CELLS, populates CELLS["WACC"]
    build_dcf_sheet(ws_dcf, CELLS)
    build_sensitivity(ws_sens, CELLS)

    out = os.path.join(os.getcwd(), filename)
    try:
        wb.save(out)
    except PermissionError:
        print(f"\n  ERROR: Could not save — {filename} is open in Excel.")
        print("  Close the file and run the command again.")
        sys.exit(1)
    print(f"Saved: {out}")
    return out


# ═══════════════════════════════════════════════════════════════════════════════
# GUIDE
# ═══════════════════════════════════════════════════════════════════════════════
def build_guide(ws):
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 62
    ws.merge_cells("B1:C1")
    sc(ws, "B1", "Professional DCF Valuation Model", bold=True, size=16,
       fc="FFFFFF", bg=NAVY, al=L)
    row_h(ws, 1, 32)
    ws.merge_cells("B2:C2")
    sc(ws, "B2", f"  Built {datetime.date.today()}  |  Gold cells = inputs  |  White cells = formulas",
       size=9, italic=True, fc="FFFFFF", bg=DARK_BLUE, al=L)
    row_h(ws, 2, 16)

    steps = [
        ("Stock Search",    "Enter ticker in B5. Microsoft 365: Data > Stocks to link live data.\n"
                            "Or run:  python build_dcf.py AAPL  to build and auto-fill in one step."),
        ("Assumptions",     "Edit all gold cells. Revenue growth, margins, WACC inputs, terminal value."),
        ("Income Statement","Historical (grey) + 10-year projection. FCF computed automatically."),
        ("WACC",            "Cost of equity (CAPM), cost of debt, capital structure -> WACC."),
        ("DCF Valuation",   "Discounted FCFs + terminal value -> Enterprise Value -> Implied Price."),
        ("Sensitivity",     "Implied price vs WACC x terminal growth rate. Green = upside."),
    ]
    r = 4
    for tab, desc in steps:
        row_h(ws, r, 14)
        ws.merge_cells(f"B{r}:C{r}")
        sc(ws, f"B{r}", f"  [{tab}]  {desc.splitlines()[0]}", bold=(r==4), size=9,
           bg=LIGHT_BLUE if r > 4 else MED_BLUE, fc="000000" if r > 4 else "FFFFFF",
           al=L, bdr=thin_border)
        for extra in desc.splitlines()[1:]:
            r += 1
            row_h(ws, r, 13)
            ws.merge_cells(f"B{r}:C{r}")
            sc(ws, f"B{r}", f"       {extra}", size=9, italic=True,
               bg=LIGHT_BLUE, al=L, bdr=thin_border)
        r += 1


# ═══════════════════════════════════════════════════════════════════════════════
# STOCK SEARCH
# ═══════════════════════════════════════════════════════════════════════════════
def build_stock(ws, CELLS):
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 18

    ws.merge_cells("B1:E1")
    sc(ws, "B1", "STOCK SEARCH  -  Data Input Center", bold=True, size=14,
       fc="FFFFFF", bg=NAVY, al=C)
    row_h(ws, 1, 30)
    ws.merge_cells("B2:E2")
    sc(ws, "B2",
       "  Enter ticker in B5 below. Microsoft 365: select B5, then Data > Stocks to link live data."
       "  Or run update_dcf.py to auto-populate.",
       size=9, italic=True, fc="404040", bg=LIGHT_BLUE, al=L)
    row_h(ws, 2, 14)

    # Header row
    r = 4
    for c, lbl_t in [("B","TICKER"), ("C","PRICE ($)"), ("D","COMPANY"), ("E","LAST UPDATED")]:
        hdr_cell(ws, f"{c}{r}", lbl_t, bg=DARK_BLUE)
    row_h(ws, r, 16)

    # Data row
    r = 5
    inp(ws, "B5", "AAPL",  "@")
    inp(ws, "C5", 175.50,  FMT_USD2)
    inp(ws, "D5", "Apple Inc.", "@")
    inp(ws, "E5", str(datetime.date.today()), "@")
    row_h(ws, r, 18)

    CELLS["SS_TICKER"]   = "'Stock Search'!B5"
    CELLS["SS_PRICE"]    = "'Stock Search'!C5"
    CELLS["SS_NAME"]     = "'Stock Search'!D5"

    # Key metrics table
    metrics = [
        ("PRICE & RANGE",     [
            ("Current Price",    "C", 175.50,     FMT_USD2),
            ("52-Week High",     "D",  198.23,    FMT_USD2),
            ("52-Week Low",      "E",  124.17,    FMT_USD2),
        ]),
        ("SHARES & MARKET CAP", [
            ("Market Cap ($M)",  "C", 2_750_000,  FMT_USD),
            ("Shares Out. (M)",  "D",    15_550,  FMT_SHR),
            ("Float %",          "E",     0.999,  FMT_PCT),
        ]),
        ("INCOME (TTM, $M)",  [
            ("Revenue",          "C", 385_000,    FMT_USD),
            ("EBITDA",           "D", 130_000,    FMT_USD),
            ("Net Income",       "E",  99_800,    FMT_USD),
        ]),
        ("BALANCE SHEET ($M)", [
            ("Total Debt",       "C", 109_000,    FMT_USD),
            ("Cash",             "D",  65_000,    FMT_USD),
            ("Net Debt",         "E",  "=C16-D16", FMT_USD),
        ]),
        ("VALUATION",         [
            ("P/E (TTM)",        "C",   28.5,     FMT_MULT),
            ("EV/EBITDA",        "D",   22.1,     FMT_MULT),
            ("Beta",             "E",    1.20,    "0.00"),
        ]),
    ]
    r = 7
    for group, fields in metrics:
        sec(ws, r, 2, 5, group)
        row_h(ws, r, 15)
        r += 1
        row_h(ws, r, 17)
        for col, val, fmt in [(c, v, f) for _, c, v, f in fields]:
            lbl_val = [lbl_t for lbl_t, c, v, f in fields if c == col][0]
            lbl(ws, f"B{r}", f"  {lbl_val}")
        # Rewrite per field
        for lbl_t, col, val, fmt in fields:
            lbl(ws, f"B{r}", f"  {fields[0][0]}")
            break
        r_save = r
        ws.merge_cells(f"B{r}:B{r}")  # ensure not merged
        # Build labels in col B and values in C/D/E
        r = r_save
        sc(ws, f"B{r}", f"  {fields[0][0]} / {fields[1][0]} / {fields[2][0]}",
           size=8, italic=True, bg=GRAY_LIGHT, al=L, bdr=thin_border)
        for lbl_t, col, val, fmt in fields:
            if isinstance(val, str) and val.startswith("="):
                cal(ws, f"{col}{r}", val, fmt)
            else:
                inp(ws, f"{col}{r}", val, fmt)
        r += 1

    # Store key cell addresses
    CELLS["SS_REV"]   = f"'Stock Search'!C13"  # Revenue TTM
    CELLS["SS_EBITDA"]= f"'Stock Search'!D13"  # EBITDA TTM
    CELLS["SS_DEBT"]  = f"'Stock Search'!C16"  # Total Debt
    CELLS["SS_CASH"]  = f"'Stock Search'!D16"  # Cash
    CELLS["SS_SHARES"]= f"'Stock Search'!D10"  # Shares Outstanding

    # M365 Stocks instructions
    r += 1
    sec(ws, r, 2, 5, "MICROSOFT 365 STOCKS DATA TYPE - INSTRUCTIONS")
    row_h(ws, r, 15)
    steps = [
        "1. Type ticker in B5  (e.g. AAPL, MSFT, GOOGL, NVDA)",
        "2. With B5 selected: click  Data  tab  >  Stocks  button in ribbon",
        "3. Excel converts the text to a linked Stocks data type (shows stock icon)",
        "4. Click the icon to browse available fields, or type  =B5.Price  to extract any field",
        "5. STOCKHISTORY(B5, start_date, end_date)  returns historical price table",
        "",
        "PYTHON AUTO-FILL (works with any Excel version):",
        "   pip install yfinance openpyxl",
        "   python build_dcf.py AAPL   (build + fill in one step)",
        "   python update_dcf.py AAPL  (refresh an existing model)",
    ]
    for step in steps:
        r += 1
        row_h(ws, r, 13)
        ws.merge_cells(f"B{r}:E{r}")
        sc(ws, f"B{r}", f"  {step}", size=9,
           italic=step.startswith("   "),
           bg=LIGHT_BLUE if step.startswith("PYTHON") else GRAY_LIGHT,
           al=L, bdr=thin_border)


# ═══════════════════════════════════════════════════════════════════════════════
# ASSUMPTIONS  (returns key cell addresses in CELLS dict)
# ═══════════════════════════════════════════════════════════════════════════════
def build_assum(ws, CELLS):
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 32
    ws.column_dimensions["E"].width = 16

    ws.merge_cells("B1:E1")
    sc(ws, "B1", "MODEL ASSUMPTIONS  -  Edit Gold Cells Only", bold=True, size=13,
       fc="FFFFFF", bg=NAVY, al=C)
    row_h(ws, 1, 26)
    ws.merge_cells("B2:E2")
    sc(ws, "B2", "  All dollar amounts in USD millions. Percentages as decimals (10% = 0.10).",
       size=9, italic=True, fc="404040", bg=LIGHT_BLUE, al=L)
    row_h(ws, 2, 13)

    def asec(row, label):
        ws.merge_cells(f"B{row}:E{row}")
        sc(ws, f"B{row}", f"  {label}", bold=True, size=9, fc="FFFFFF",
           bg=DARK_BLUE, al=L, bdr=thin_border)
        row_h(ws, row, 17)

    def arow(r, l1, v1, f1, l2=None, v2=None, f2=None):
        row_h(ws, r, 17)
        lbl(ws, f"B{r}", f"  {l1}")
        inp(ws, f"C{r}", v1, f1)
        if l2:
            lbl(ws, f"D{r}", f"  {l2}")
            inp(ws, f"E{r}", v2, f2)
        else:
            sc(ws, f"D{r}", "", bg=GRAY_LIGHT, bdr=thin_border)
            sc(ws, f"E{r}", "", bg=GRAY_LIGHT, bdr=thin_border)

    r = 4
    asec(r, "COMPANY")
    r += 1; arow(r, "Company Name",         "Apple Inc.",  "@",      "Ticker",            "AAPL",    "@")
    r += 1; arow(r, "Fiscal Year End",      "September",   "@",      "Currency",          "USD",     "@")
    r += 1; arow(r, "Current Price ($)",    175.50,        FMT_USD2, "Shares Out. (M)",   15_550,    FMT_SHR)

    CELLS["A_PRICE"]  = f"Assumptions!C{r}"
    CELLS["A_SHARES"] = f"Assumptions!E{r}"

    r += 1; arow(r, "Total Debt ($M)",      109_000,       FMT_USD,  "Cash ($M)",         65_000,    FMT_USD)

    CELLS["A_DEBT"]   = f"Assumptions!C{r}"
    CELLS["A_CASH"]   = f"Assumptions!E{r}"

    r += 1
    asec(r, "REVENUE GROWTH  (10-year projection ramp)")
    r += 1; arow(r, "Y1 Revenue Growth",    0.070,  FMT_PCT2, "Y2 Revenue Growth",  0.065, FMT_PCT2)
    CELLS["A_G1"] = f"Assumptions!C{r}"
    CELLS["A_G2"] = f"Assumptions!E{r}"
    r += 1; arow(r, "Y3 Revenue Growth",    0.060,  FMT_PCT2, "Y4 Revenue Growth",  0.055, FMT_PCT2)
    CELLS["A_G3"] = f"Assumptions!C{r}"
    CELLS["A_G4"] = f"Assumptions!E{r}"
    r += 1; arow(r, "Y5 Revenue Growth",    0.050,  FMT_PCT2, "Y6 Revenue Growth",  0.045, FMT_PCT2)
    CELLS["A_G5"] = f"Assumptions!C{r}"
    CELLS["A_G6"] = f"Assumptions!E{r}"
    r += 1; arow(r, "Y7 Revenue Growth",    0.040,  FMT_PCT2, "Y8 Revenue Growth",  0.035, FMT_PCT2)
    CELLS["A_G7"] = f"Assumptions!C{r}"
    CELLS["A_G8"] = f"Assumptions!E{r}"
    r += 1; arow(r, "Y9 Revenue Growth",    0.030,  FMT_PCT2, "Y10 Revenue Growth", 0.030, FMT_PCT2)
    CELLS["A_G9"]  = f"Assumptions!C{r}"
    CELLS["A_G10"] = f"Assumptions!E{r}"
    CELLS["A_GROWTH"] = [CELLS[f"A_G{i}"] for i in range(1, 11)]

    r += 1
    asec(r, "MARGIN ASSUMPTIONS")
    r += 1; arow(r, "Gross Margin %",       0.435, FMT_PCT2, "EBITDA Margin %",     0.335, FMT_PCT2)
    CELLS["A_GM"]    = f"Assumptions!C{r}"
    CELLS["A_EBITDA_M"] = f"Assumptions!E{r}"
    r += 1; arow(r, "D&A (% of Revenue)",   0.025, FMT_PCT2, "CapEx (% of Revenue)",0.028, FMT_PCT2)
    CELLS["A_DA"]    = f"Assumptions!C{r}"
    CELLS["A_CAPEX"] = f"Assumptions!E{r}"
    r += 1; arow(r, "Net Working Cap (% Rev)",0.035,FMT_PCT2,"Tax Rate",            0.155, FMT_PCT2)
    CELLS["A_NWC"]   = f"Assumptions!C{r}"
    CELLS["A_TAX"]   = f"Assumptions!E{r}"
    r += 1; arow(r, "SBC (% of Revenue)",   0.025, FMT_PCT2, "", None, None)
    CELLS["A_SBC"]   = f"Assumptions!C{r}"

    r += 1
    asec(r, "WACC INPUTS")
    r += 1; arow(r, "Risk-Free Rate (Rf)",  0.043, FMT_PCT2, "Equity Risk Premium", 0.055, FMT_PCT2)
    CELLS["A_RF"]    = f"Assumptions!C{r}"
    CELLS["A_ERP"]   = f"Assumptions!E{r}"
    r += 1; arow(r, "Beta (Levered)",        1.20,  "0.00",  "Pre-Tax Cost of Debt",0.038, FMT_PCT2)
    CELLS["A_BETA"]  = f"Assumptions!C{r}"
    CELLS["A_KD"]    = f"Assumptions!E{r}"
    r += 1; arow(r, "Debt / Total Capital", 0.18,  FMT_PCT2, "Equity / Capital",    0.82,  FMT_PCT2)
    CELLS["A_WD"]    = f"Assumptions!C{r}"
    CELLS["A_WE"]    = f"Assumptions!E{r}"
    r += 1
    row_h(ws, r, 17)
    lbl(ws, f"B{r}", "  WACC (from WACC sheet)", bold=True, bg=LIGHT_BLUE)
    sc(ws, f"C{r}", "=WACC!E_WACC",  # placeholder; will be overwritten after WACC built
       fmt=FMT_PCT2, bold=True, bg=LIGHT_BLUE, al=R, bdr=thin_border)
    sc(ws, f"D{r}", "", bg=LIGHT_BLUE, bdr=thin_border)
    sc(ws, f"E{r}", "", bg=LIGHT_BLUE, bdr=thin_border)
    CELLS["A_WACC_ROW"] = r  # store row so WACC sheet can fix the formula
    CELLS["A_WACC_CELL"] = f"Assumptions!C{r}"

    r += 1
    asec(r, "TERMINAL VALUE")
    r += 1; arow(r, "Terminal Growth Rate", 0.025, FMT_PCT2, "Exit EBITDA Multiple",18.0,  FMT_MULT)
    CELLS["A_TVG"]   = f"Assumptions!C{r}"
    CELLS["A_TVMUL"] = f"Assumptions!E{r}"

    r += 1
    asec(r, "HISTORICAL REVENUE (for Y0 base)")
    r += 1; arow(r, "Base Revenue ($M)  [LTM/Y0]", 385_000, FMT_USD,
                    "FY-1 Revenue ($M)",           365_000, FMT_USD)
    CELLS["A_REV0"]  = f"Assumptions!C{r}"
    CELLS["A_REV_1"] = f"Assumptions!E{r}"
    r += 1; arow(r, "FY-2 Revenue ($M)",   348_000, FMT_USD, "", None, None)
    CELLS["A_REV_2"] = f"Assumptions!C{r}"

    r += 2
    asec(r, "INVESTMENT THESIS (free text)")
    for note in ["Company and thesis notes here.", "Key risks:", "Key catalysts:"]:
        r += 1
        row_h(ws, r, 15)
        ws.merge_cells(f"B{r}:E{r}")
        inp(ws, f"B{r}", note, "@")


# ═══════════════════════════════════════════════════════════════════════════════
# INCOME STATEMENT  (10-year projection + FCF)
# ═══════════════════════════════════════════════════════════════════════════════
def build_is(ws, CELLS):
    A = CELLS  # shorthand
    current_year = datetime.date.today().year
    PROJ_YEARS = list(range(current_year + 1, current_year + 11))  # Y1..Y10

    # Column layout:
    # B = label, C = FY-2, D = FY-1, E = LTM/Y0, F..O = Y1..Y10 (13 data cols total)
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 38
    for ci in range(3, 16):
        ws.column_dimensions[get_column_letter(ci)].width = 12

    # Title
    ws.merge_cells("B1:O1")
    sc(ws, "B1", "INCOME STATEMENT & FREE CASH FLOW  (USD $M)", bold=True, size=12,
       fc="FFFFFF", bg=NAVY, al=C)
    row_h(ws, 1, 26)

    # Year group header
    r = 3
    sc(ws, f"B{r}", "", bg=NAVY)
    ws.merge_cells(f"C{r}:E{r}")
    sc(ws, f"C{r}", "HISTORICAL", bold=True, fc="FFFFFF", bg=MED_BLUE, al=C)
    ws.merge_cells(f"F{r}:O{r}")
    sc(ws, f"F{r}", "PROJECTED  (10 years)", bold=True, fc="FFFFFF", bg=DARK_BLUE, al=C)
    row_h(ws, r, 13)

    r = 4
    sc(ws, f"B{r}", "  Metric (USD $M)", bold=True, size=9, fc="FFFFFF", bg=NAVY, al=L)
    GROWTH_REFS = A["A_GROWTH"]  # list of 10 Assumptions cell references
    HIST_BKGS   = [MED_BLUE, MED_BLUE, MED_BLUE]
    PROJ_BKGS   = [DARK_BLUE] * 10
    ALL_YEARS   = [current_year - 2, current_year - 1, current_year] + PROJ_YEARS
    ALL_BGS     = HIST_BKGS + PROJ_BKGS
    for ci, (yr, bg) in enumerate(zip(ALL_YEARS, ALL_BGS)):
        col = get_column_letter(3 + ci)
        label = f"FY{yr}" + ("A" if ci < 3 else "E")
        hdr_cell(ws, f"{col}{r}", label, bg=bg, size=9)
    row_h(ws, r, 16)

    # Helper to get column letter for year index 0..12
    def col(i): return get_column_letter(3 + i)  # i=0->C, 1->D, ..., 12->O

    # ── REVENUE ──────────────────────────────────────────────────────────────
    r += 1; sec(ws, r, 2, 15, "REVENUE")
    row_h(ws, r, 14)

    # Revenue growth % row
    r += 1
    row_h(ws, r, 14)
    lbl(ws, f"B{r}", "  Revenue Growth %", italic=True)
    sc(ws, f"C{r}", "—", size=8, bg=GRAY_LIGHT, al=C, bdr=thin_border)
    for ci in range(1, 3):
        cal(ws, f"{col(ci)}{r}",
            f"=({col(ci)}REV-{col(ci-1)}REV)/{col(ci-1)}REV",  # placeholder; set below
            FMT_PCT2)
    for gi, gref in enumerate(GROWTH_REFS):
        cal(ws, f"{col(3+gi)}{r}", f"={gref}", FMT_PCT2)

    REV_GR_ROW = r

    # Revenue $ row
    r += 1
    row_h(ws, r, 18)
    lbl(ws, f"B{r}", "  Revenue", bold=True, bg=LIGHT_BLUE)
    # Historical: 3 input cells from Assumptions
    inp(ws, f"C{r}", f"={A['A_REV_2']}", FMT_USD)
    inp(ws, f"D{r}", f"={A['A_REV_1']}", FMT_USD)
    inp(ws, f"E{r}", f"={A['A_REV0']}",  FMT_USD)
    # Projected: prior * (1 + growth)
    for gi, gref in enumerate(GROWTH_REFS):
        prev_col = col(2 + gi)
        cur_col  = col(3 + gi)
        cal(ws, f"{cur_col}{r}", f"={prev_col}{r}*(1+{gref})", FMT_USD, bold=True)

    REV_ROW = r

    # Fix revenue growth formula using actual row
    for ci in range(1, 3):
        ws[f"{col(ci)}{REV_GR_ROW}"].value = \
            f"=({col(ci)}{REV_ROW}-{col(ci-1)}{REV_ROW})/{col(ci-1)}{REV_ROW}"

    CELLS["IS_REV_ROW"] = REV_ROW

    # ── GROSS PROFIT ─────────────────────────────────────────────────────────
    r += 1; sec(ws, r, 2, 15, "GROSS PROFIT")
    row_h(ws, r, 14)

    r += 1
    row_h(ws, r, 14)
    lbl(ws, f"B{r}", "  Gross Margin %", italic=True)
    for ci in [0, 1, 2]:
        inp(ws, f"{col(ci)}{r}", [0.432, 0.433, 0.435][ci], FMT_PCT2)
    for gi in range(10):
        cal(ws, f"{col(3+gi)}{r}", f"={A['A_GM']}", FMT_PCT2)
    GM_ROW = r

    r += 1
    row_h(ws, r, 18)
    lbl(ws, f"B{r}", "  Gross Profit", bold=True, bg=LIGHT_BLUE)
    for ci in range(13):
        cal(ws, f"{col(ci)}{r}", f"={col(ci)}{REV_ROW}*{col(ci)}{GM_ROW}", FMT_USD, bold=True)
    GP_ROW = r

    # ── EBITDA ───────────────────────────────────────────────────────────────
    r += 1; sec(ws, r, 2, 15, "OPERATING EXPENSES -> EBITDA")
    row_h(ws, r, 14)

    r += 1
    row_h(ws, r, 14)
    lbl(ws, f"B{r}", "  EBITDA Margin %", italic=True)
    for ci in [0, 1, 2]:
        inp(ws, f"{col(ci)}{r}", [0.327, 0.330, 0.335][ci], FMT_PCT2)
    for gi in range(10):
        cal(ws, f"{col(3+gi)}{r}", f"={A['A_EBITDA_M']}", FMT_PCT2)
    EBITDA_M_ROW = r

    r += 1
    row_h(ws, r, 18)
    lbl(ws, f"B{r}", "  EBITDA", bold=True, bg=LIGHT_BLUE)
    for ci in range(13):
        cal(ws, f"{col(ci)}{r}", f"={col(ci)}{REV_ROW}*{col(ci)}{EBITDA_M_ROW}", FMT_USD, bold=True)
    EBITDA_ROW = r
    CELLS["IS_EBITDA_ROW"] = EBITDA_ROW

    r += 1
    row_h(ws, r, 14)
    lbl(ws, f"B{r}", "  OpEx (= Gross Profit - EBITDA)", italic=True)
    for ci in range(13):
        cal(ws, f"{col(ci)}{r}", f"={col(ci)}{GP_ROW}-{col(ci)}{EBITDA_ROW}", FMT_USD)
    OPEX_ROW = r

    # ── D&A / EBIT ───────────────────────────────────────────────────────────
    r += 1; sec(ws, r, 2, 15, "D&A  ->  EBIT")
    row_h(ws, r, 14)

    r += 1
    row_h(ws, r, 14)
    lbl(ws, f"B{r}", "  D&A % of Revenue", italic=True)
    for ci in [0, 1, 2]:
        inp(ws, f"{col(ci)}{r}", 0.025, FMT_PCT2)
    for gi in range(10):
        cal(ws, f"{col(3+gi)}{r}", f"={A['A_DA']}", FMT_PCT2)
    DA_PCT_ROW = r

    r += 1
    row_h(ws, r, 14)
    lbl(ws, f"B{r}", "  D&A ($M)", italic=True)
    for ci in range(13):
        cal(ws, f"{col(ci)}{r}", f"={col(ci)}{DA_PCT_ROW}*{col(ci)}{REV_ROW}", FMT_USD)
    DA_ROW = r

    r += 1
    row_h(ws, r, 18)
    lbl(ws, f"B{r}", "  EBIT (Operating Income)", bold=True, bg=LIGHT_BLUE)
    for ci in range(13):
        cal(ws, f"{col(ci)}{r}", f"={col(ci)}{EBITDA_ROW}-{col(ci)}{DA_ROW}", FMT_USD, bold=True)
    EBIT_ROW = r

    r += 1
    row_h(ws, r, 14)
    lbl(ws, f"B{r}", "  EBIT Margin %", italic=True)
    for ci in range(13):
        cal(ws, f"{col(ci)}{r}",
            f"=IF({col(ci)}{REV_ROW}=0,0,{col(ci)}{EBIT_ROW}/{col(ci)}{REV_ROW})",
            FMT_PCT2)

    # ── NOPAT ────────────────────────────────────────────────────────────────
    r += 1; sec(ws, r, 2, 15, "NOPAT  (Net Operating Profit After Tax)")
    row_h(ws, r, 14)

    r += 1
    row_h(ws, r, 14)
    lbl(ws, f"B{r}", "  Tax Rate %", italic=True)
    for ci in [0, 1, 2]:
        inp(ws, f"{col(ci)}{r}", [0.140, 0.148, 0.155][ci], FMT_PCT2)
    for gi in range(10):
        cal(ws, f"{col(3+gi)}{r}", f"={A['A_TAX']}", FMT_PCT2)
    TAX_ROW = r

    r += 1
    row_h(ws, r, 18)
    lbl(ws, f"B{r}", "  NOPAT = EBIT x (1 - Tax Rate)", bold=True, bg=LIGHT_BLUE)
    for ci in range(13):
        cal(ws, f"{col(ci)}{r}",
            f"={col(ci)}{EBIT_ROW}*(1-{col(ci)}{TAX_ROW})", FMT_USD, bold=True)
    NOPAT_ROW = r

    # ── FREE CASH FLOW ────────────────────────────────────────────────────────
    r += 1; sec(ws, r, 2, 15, "FREE CASH FLOW TO FIRM (FCFF)")
    row_h(ws, r, 14)

    r += 1
    row_h(ws, r, 14)
    lbl(ws, f"B{r}", "  (+) D&A add-back ($M)", italic=True)
    for ci in range(13):
        cal(ws, f"{col(ci)}{r}", f"={col(ci)}{DA_ROW}", FMT_USD)
    DA_BACK_ROW = r

    r += 1
    row_h(ws, r, 14)
    lbl(ws, f"B{r}", "  (-) CapEx % of Revenue", italic=True)
    for ci in [0, 1, 2]:
        inp(ws, f"{col(ci)}{r}", 0.028, FMT_PCT2)
    for gi in range(10):
        cal(ws, f"{col(3+gi)}{r}", f"={A['A_CAPEX']}", FMT_PCT2)
    CAPEX_PCT_ROW = r

    r += 1
    row_h(ws, r, 14)
    lbl(ws, f"B{r}", "  (-) CapEx ($M)", italic=True)
    for ci in range(13):
        cal(ws, f"{col(ci)}{r}",
            f"={col(ci)}{CAPEX_PCT_ROW}*{col(ci)}{REV_ROW}", FMT_USD)
    CAPEX_ROW = r

    r += 1
    row_h(ws, r, 14)
    lbl(ws, f"B{r}", "  (+/-) Change in Net Working Capital ($M)", italic=True)
    # Hist: manual inputs
    inp(ws, f"C{r}", -3_200, FMT_USD)
    inp(ws, f"D{r}", -3_800, FMT_USD)
    # LTM: use NWC % * Rev
    cal(ws, f"E{r}", f"=-{A['A_NWC']}*E{REV_ROW}", FMT_USD)
    for gi in range(10):
        cur = col(3 + gi)
        prv = col(2 + gi)
        cal(ws, f"{cur}{r}",
            f"=-({A['A_NWC']}*{cur}{REV_ROW}-{A['A_NWC']}*{prv}{REV_ROW})",
            FMT_USD)
    NWC_ROW = r

    r += 1
    row_h(ws, r, 14)
    lbl(ws, f"B{r}", "  (+) Stock-Based Comp ($M)", italic=True)
    inp(ws, f"C{r}", 9_200, FMT_USD)
    inp(ws, f"D{r}", 9_700, FMT_USD)
    for gi in range(-1, 10):
        ci = 2 + gi if gi >= 0 else 2
        c_ref = col(3 + gi) if gi >= 0 else "E"
        cal(ws, f"{c_ref}{r}",
            f"={A['A_SBC']}*{c_ref}{REV_ROW}", FMT_USD)
    SBC_ROW = r

    r += 1
    row_h(ws, r, 20)
    sc(ws, f"B{r}", "  FCFF  (Free Cash Flow to Firm)", bold=True, size=11,
       fc="FFFFFF", bg=NAVY, al=L, bdr=thin_border)
    for ci in range(13):
        cal(ws, f"{col(ci)}{r}",
            f"={col(ci)}{NOPAT_ROW}"
            f"+{col(ci)}{DA_BACK_ROW}"
            f"-{col(ci)}{CAPEX_ROW}"
            f"+{col(ci)}{NWC_ROW}"
            f"+{col(ci)}{SBC_ROW}",
            FMT_USD, bold=True)
        ws[f"{col(ci)}{r}"].fill = Fill(DARK_BLUE)
        ws[f"{col(ci)}{r}"].font = F(bold=True, color="FFFFFF")
    FCFF_ROW = r
    CELLS["IS_FCFF_ROW"] = FCFF_ROW

    r += 1
    row_h(ws, r, 14)
    lbl(ws, f"B{r}", "  FCF Margin %", italic=True)
    for ci in range(13):
        cal(ws, f"{col(ci)}{r}",
            f"=IF({col(ci)}{REV_ROW}=0,0,{col(ci)}{FCFF_ROW}/{col(ci)}{REV_ROW})",
            FMT_PCT2)

    # Store IS sheet references for DCF
    # Projected FCF runs from col F (ci=3) to O (ci=12)
    CELLS["IS_FCFF_PROJ"] = (
        f"'Income Statement'!F{FCFF_ROW}",
        f"'Income Statement'!O{FCFF_ROW}",
        FCFF_ROW
    )
    CELLS["IS_EBITDA_Y10"] = f"'Income Statement'!O{EBITDA_ROW}"

    # ── FCF Chart ─────────────────────────────────────────────────────────────
    r += 2
    row_h(ws, r, 16)
    ws.merge_cells(f"B{r}:O{r}")
    sc(ws, f"B{r}", "  10-YEAR FCF PROJECTION (from model above)", bold=True, size=9,
       fc="FFFFFF", bg=DARK_BLUE, al=L, bdr=thin_border)

    chart = BarChart()
    chart.type = "col"
    chart.title = "Projected FCFF  (USD $M)"
    chart.y_axis.title = "FCFF ($M)"
    chart.style = 2
    chart.height = 10
    chart.width  = 24
    fcf_ref = Reference(ws, min_col=6, min_row=FCFF_ROW, max_col=15, max_row=FCFF_ROW)
    yr_ref  = Reference(ws, min_col=6, min_row=4,         max_col=15, max_row=4)
    chart.add_data(fcf_ref)
    chart.set_categories(yr_ref)
    if chart.series:
        chart.series[0].graphicalProperties.solidFill = "2E4170"
    ws.add_chart(chart, f"B{r+1}")


# ═══════════════════════════════════════════════════════════════════════════════
# WACC SHEET
# ═══════════════════════════════════════════════════════════════════════════════
def build_wacc(ws, CELLS):
    A = CELLS
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 18

    ws.merge_cells("B1:E1")
    sc(ws, "B1", "WACC  -  Weighted Average Cost of Capital", bold=True, size=13,
       fc="FFFFFF", bg=NAVY, al=C)
    row_h(ws, 1, 26)
    ws.merge_cells("B2:E2")
    sc(ws, "B2",
       "  WACC = Ke x We  +  Kd x (1-t) x Wd    |    Sources: Assumptions sheet",
       size=9, italic=True, fc="404040", bg=LIGHT_BLUE, al=L)
    row_h(ws, 2, 13)

    def wsec(row, label):
        ws.merge_cells(f"B{row}:E{row}")
        sc(ws, f"B{row}", f"  {label}", bold=True, size=9, fc="FFFFFF",
           bg=DARK_BLUE, al=L, bdr=thin_border)
        row_h(ws, row, 16)

    def wrow(r, l1, v1, f1, l2=None, v2=None, f2=None, editable=False):
        row_h(ws, r, 17)
        lbl(ws, f"B{r}", f"  {l1}")
        fn = inp if editable else cal
        fn(ws, f"C{r}", v1, f1)
        if l2:
            lbl(ws, f"D{r}", f"  {l2}")
            fn(ws, f"E{r}", v2, f2)

    r = 4
    wsec(r, "COST OF EQUITY  -  CAPM:  Ke = Rf + beta x ERP")
    r += 1; wrow(r, "Risk-Free Rate (Rf)", f"={A['A_RF']}",   FMT_PCT2,
                    "Equity Risk Premium",  f"={A['A_ERP']}",  FMT_PCT2)
    r += 1; wrow(r, "Beta (Levered)",       f"={A['A_BETA']}", "0.00",  editable=False)

    r += 1
    row_h(ws, r, 20)
    ws.merge_cells(f"B{r}:D{r}")
    sc(ws, f"B{r}", "  COST OF EQUITY (Ke)", bold=True, size=11, fc="FFFFFF",
       bg=DARK_BLUE, al=L, bdr=thin_border)
    ke_formula = f"={A['A_RF']}+{A['A_BETA']}*{A['A_ERP']}"
    cal(ws, f"E{r}", ke_formula, FMT_PCT2, bold=True)
    ws[f"E{r}"].fill = Fill(MED_BLUE)
    ws[f"E{r}"].font = F(bold=True, size=11, color="FFFFFF")
    KE_CELL = f"WACC!E{r}"

    r += 1
    wsec(r, "COST OF DEBT  -  After Tax:  Kd_at = Kd x (1 - Tax Rate)")
    r += 1; wrow(r, "Pre-Tax Cost of Debt (Kd)", f"={A['A_KD']}", FMT_PCT2,
                    "Tax Rate",                   f"={A['A_TAX']}", FMT_PCT2)
    KD_PRE_ROW = r

    r += 1
    row_h(ws, r, 20)
    ws.merge_cells(f"B{r}:D{r}")
    sc(ws, f"B{r}", "  AFTER-TAX COST OF DEBT (Kd)", bold=True, size=11, fc="FFFFFF",
       bg=DARK_BLUE, al=L, bdr=thin_border)
    cal(ws, f"E{r}", f"=C{KD_PRE_ROW}*(1-E{KD_PRE_ROW})", FMT_PCT2, bold=True)
    ws[f"E{r}"].fill = Fill(MED_BLUE)
    ws[f"E{r}"].font = F(bold=True, size=11, color="FFFFFF")
    KD_CELL = f"WACC!E{r}"

    r += 1
    wsec(r, "CAPITAL STRUCTURE WEIGHTS")
    r += 1; wrow(r, "Equity / Capital (We)", f"={A['A_WE']}", FMT_PCT2,
                    "Debt / Capital (Wd)",   f"={A['A_WD']}", FMT_PCT2)
    WE_ROW = r; WE_COL = "C"; WD_COL = "E"
    r += 1
    row_h(ws, r, 15)
    lbl(ws, f"B{r}", "  Check: We + Wd should = 1.0", italic=True)
    cal(ws, f"C{r}", f"=C{WE_ROW}+E{WE_ROW}", FMT_PCT2)
    lbl(ws, f"D{r}", "  Status:", italic=True)
    cal(ws, f"E{r}", f'=IF(ABS(C{WE_ROW}+E{WE_ROW}-1)<0.001,"OK - sums to 100%","CHECK WEIGHTS")', "@")

    r += 1
    wsec(r, "WACC CALCULATION")
    r += 1
    row_h(ws, r, 17)
    lbl(ws, f"B{r}", "  Ke x We  (equity component)")
    cal(ws, f"C{r}", f"={KE_CELL}*C{WE_ROW}", FMT_PCT2)
    lbl(ws, f"D{r}", "  Kd(1-t) x Wd  (debt component)")
    cal(ws, f"E{r}", f"={KD_CELL}*E{WE_ROW}", FMT_PCT2)
    r += 1
    row_h(ws, r, 24)
    ws.merge_cells(f"B{r}:D{r}")
    sc(ws, f"B{r}", "  WACC  =  Ke x We  +  Kd(1-t) x Wd", bold=True, size=13,
       fc="FFFFFF", bg=NAVY, al=L, bdr=thin_border)
    cal(ws, f"E{r}", f"=C{r-1}+E{r-1}", FMT_PCT2, bold=True)
    ws[f"E{r}"].fill = Fill(NAVY)
    ws[f"E{r}"].font = F(bold=True, size=14, color="FFFFFF")
    WACC_ROW = r
    CELLS["WACC_CELL"] = f"WACC!E{r}"

    # Fix Assumptions WACC reference now that we know the row
    # We stored the row in A_WACC_ROW; find the workbook via ws.parent
    if "A_WACC_ROW" in CELLS:
        assum_ws = ws.parent["Assumptions"]
        wacc_ref_row = CELLS["A_WACC_ROW"]
        assum_ws[f"C{wacc_ref_row}"].value = f"={CELLS['WACC_CELL']}"

    r += 2
    ws.merge_cells(f"B{r}:E{r}")
    sc(ws, f"B{r}",
       "Note: WACC above flows into DCF Valuation sheet automatically. "
       "Sensitivity analysis spans +/-150bps in the Sensitivity sheet.",
       size=9, italic=True, fc="606060", bg=GRAY_LIGHT, al=L)


# ═══════════════════════════════════════════════════════════════════════════════
# DCF VALUATION SHEET
# ═══════════════════════════════════════════════════════════════════════════════
def build_dcf_sheet(ws, CELLS):
    A = CELLS
    current_year = datetime.date.today().year
    PROJ_YEARS   = list(range(current_year + 1, current_year + 11))

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 18

    ws.merge_cells("B1:E1")
    sc(ws, "B1", "DCF VALUATION  -  Enterprise & Equity Value Bridge", bold=True, size=13,
       fc="FFFFFF", bg=NAVY, al=C)
    row_h(ws, 1, 26)

    # Project FCFF rows
    FCFF_ROW = A["IS_FCFF_ROW"]
    EBITDA_Y10 = A["IS_EBITDA_Y10"]
    WACC_CELL  = A["WACC_CELL"]   # e.g. "WACC!E16"
    TVG_CELL   = A["A_TVG"]
    TVMUL_CELL = A["A_TVMUL"]

    # Build 10 FCF column formulas (IS projected cols F..O = col indices 6..15)
    PROJ_FCF_COLS = [get_column_letter(6 + i) for i in range(10)]
    # Individual IS FCFF references
    fcf_refs = [f"'Income Statement'!{c}{FCFF_ROW}" for c in PROJ_FCF_COLS]

    def dsec(row, label):
        ws.merge_cells(f"B{row}:E{row}")
        sc(ws, f"B{row}", f"  {label}", bold=True, size=9, fc="FFFFFF",
           bg=DARK_BLUE, al=L, bdr=thin_border)
        row_h(ws, row, 16)

    r = 3
    dsec(r, "DISCOUNT RATE & KEY INPUTS")
    params = [
        ("WACC",                      f"={WACC_CELL}",   FMT_PCT2),
        ("Terminal Growth Rate (g)",  f"={TVG_CELL}",    FMT_PCT2),
        ("Exit EBITDA Multiple",      f"={TVMUL_CELL}",  FMT_MULT),
        ("Net Debt ($M)",             f"={A['A_DEBT']}-{A['A_CASH']}", FMT_USD),
        ("Shares Outstanding (M)",    f"={A['A_SHARES']}", FMT_SHR),
    ]
    PARAM_CELLS = {}
    for label, formula, fmt in params:
        r += 1
        row_h(ws, r, 17)
        lbl(ws, f"B{r}", f"  {label}")
        cal(ws, f"C{r}", formula, fmt)
        sc(ws, f"D{r}", "", bg=GRAY_LIGHT, bdr=thin_border)
        sc(ws, f"E{r}", "", bg=GRAY_LIGHT, bdr=thin_border)
        PARAM_CELLS[label] = f"C{r}"

    WACC_REF   = PARAM_CELLS["WACC"]
    NETD_REF   = PARAM_CELLS["Net Debt ($M)"]
    SOUT_REF   = PARAM_CELLS["Shares Outstanding (M)"]
    TVG_REF    = PARAM_CELLS["Terminal Growth Rate (g)"]
    TVMUL_REF  = PARAM_CELLS["Exit EBITDA Multiple"]

    CELLS["DCF_WACC_REF"]  = f"'DCF Valuation'!{WACC_REF}"
    CELLS["DCF_TVG_REF"]   = f"'DCF Valuation'!{TVG_REF}"

    # FCF reference section
    r += 1
    dsec(r, "PROJECTED FCF BY YEAR  (from Income Statement)")

    # Year headers
    r += 1; row_h(ws, r, 14)
    lbl(ws, f"B{r}", "  Year")
    for i, yr in enumerate(PROJ_YEARS):
        col_let = get_column_letter(3 + i)
        if i < 5:
            ws.column_dimensions[col_let].width = 11
        sc(ws, f"{col_let}{r}", f"FY{yr}", bold=True, size=8, fc="FFFFFF",
           bg=DARK_BLUE, al=C, bdr=thin_border)
    # We only have 10 projected years -> columns C..L (3..12)
    # But layout has B=label, C..L=Y1..Y10, so we need to adjust columns
    # Use columns C..L (ci 0..9)
    for ci in range(10, 13):
        ws.column_dimensions[get_column_letter(3 + ci)].width = 1

    YR_HDR_ROW = r

    r += 1; row_h(ws, r, 17)
    lbl(ws, f"B{r}", "  FCFF ($M)", bold=True, bg=LIGHT_BLUE)
    for i, fcf_ref in enumerate(fcf_refs):
        c = get_column_letter(3 + i)
        cal(ws, f"{c}{r}", f"={fcf_ref}", FMT_USD, bold=True)
    FCFF_PROJ_ROW = r

    r += 1; row_h(ws, r, 14)
    lbl(ws, f"B{r}", "  Discount Period (t)")
    for i in range(10):
        c = get_column_letter(3 + i)
        cal(ws, f"{c}{r}", i + 1, FMT_MULT)
    PERIOD_ROW = r

    r += 1; row_h(ws, r, 14)
    lbl(ws, f"B{r}", "  Discount Factor  [1/(1+WACC)^t]")
    for i in range(10):
        c = get_column_letter(3 + i)
        p_cell = f"{get_column_letter(3+i)}{PERIOD_ROW}"
        cal(ws, f"{c}{r}", f"=1/(1+{WACC_REF})^{p_cell}", "0.0000")
    DF_ROW = r

    r += 1; row_h(ws, r, 17)
    lbl(ws, f"B{r}", "  PV of FCFF ($M)", bold=True, bg=LIGHT_BLUE)
    for i in range(10):
        c = get_column_letter(3 + i)
        cal(ws, f"{c}{r}",
            f"={c}{FCFF_PROJ_ROW}*{c}{DF_ROW}", FMT_USD, bold=True)
    PV_FCF_ROW = r

    # Sum of PVs
    r += 1; row_h(ws, r, 17)
    lbl(ws, f"B{r}", "  Sum of PV(FCFF), Years 1-10")
    cal(ws, f"C{r}", f"=SUM(C{PV_FCF_ROW}:L{PV_FCF_ROW})", FMT_USD, bold=True)
    SUM_PV_CELL = f"C{r}"

    # Terminal value
    r += 1; dsec(r, "TERMINAL VALUE")

    r += 1; row_h(ws, r, 17)
    lbl(ws, f"B{r}", "  Gordon Growth TV = FCFF_Y10 * (1+g) / (WACC - g)")
    yr10_fcf = f"L{FCFF_PROJ_ROW}"  # col L = ci=9 = Y10
    cal(ws, f"C{r}",
        f"={yr10_fcf}*(1+{TVG_REF})/({WACC_REF}-{TVG_REF})", FMT_USD)
    lbl(ws, f"D{r}", "  PV of TV (Gordon Growth)")
    cal(ws, f"E{r}", f"=C{r}/(1+{WACC_REF})^10", FMT_USD, bold=True)
    TV_GGM_ROW = r; PV_TV_GGM = f"E{r}"

    r += 1; row_h(ws, r, 17)
    lbl(ws, f"B{r}", "  Exit Multiple TV = EBITDA_Y10 x Multiple  [reference only]")
    cal(ws, f"C{r}", f"={EBITDA_Y10}*{TVMUL_REF}", FMT_USD)
    lbl(ws, f"D{r}", "  PV of TV (Exit Multiple)")
    cal(ws, f"E{r}", f"=C{r}/(1+{WACC_REF})^10", FMT_USD)

    # Valuation bridge
    r += 1; dsec(r, "VALUATION BRIDGE")

    r += 1; row_h(ws, r, 17)
    lbl(ws, f"B{r}", "  (+) Sum of PV(FCF) - Years 1-10")
    cal(ws, f"C{r}", f"={SUM_PV_CELL}", FMT_USD, bold=True)
    SUM_PV_REF = f"C{r}"

    r += 1; row_h(ws, r, 17)
    lbl(ws, f"B{r}", "  (+) PV of Terminal Value (Gordon Growth - PRIMARY)", bold=True, bg=LIGHT_BLUE)
    cal(ws, f"C{r}", f"={PV_TV_GGM}", FMT_USD, bold=True)
    PV_TV_REF = f"C{r}"

    r += 1; row_h(ws, r, 22)
    sc(ws, f"B{r}", "  ENTERPRISE VALUE", bold=True, size=12, fc="FFFFFF",
       bg=DARK_BLUE, al=L, bdr=thin_border)
    cal(ws, f"C{r}", f"={SUM_PV_REF}+{PV_TV_REF}", FMT_USD, bold=True)
    ws[f"C{r}"].fill = Fill(DARK_BLUE); ws[f"C{r}"].font = F(bold=True, size=12, color="FFFFFF")
    EV_CELL = f"C{r}"

    r += 1; row_h(ws, r, 17)
    lbl(ws, f"B{r}", "  (-) Net Debt ($M)")
    cal(ws, f"C{r}", f"={NETD_REF}", FMT_USD)
    ND_REF = f"C{r}"

    r += 1; row_h(ws, r, 22)
    sc(ws, f"B{r}", "  EQUITY VALUE", bold=True, size=12, fc="FFFFFF",
       bg=DARK_BLUE, al=L, bdr=thin_border)
    cal(ws, f"C{r}", f"={EV_CELL}-{ND_REF}", FMT_USD, bold=True)
    ws[f"C{r}"].fill = Fill(DARK_BLUE); ws[f"C{r}"].font = F(bold=True, size=12, color="FFFFFF")
    EQ_CELL = f"C{r}"

    r += 1; row_h(ws, r, 17)
    lbl(ws, f"B{r}", "  Shares Outstanding (M)")
    cal(ws, f"C{r}", f"={SOUT_REF}", FMT_SHR)
    SOUT_REF2 = f"C{r}"

    r += 1; row_h(ws, r, 28)
    sc(ws, f"B{r}", "  IMPLIED SHARE PRICE (DCF)", bold=True, size=14,
       fc="FFFFFF", bg=NAVY, al=L, bdr=thin_border)
    cal(ws, f"C{r}", f"={EQ_CELL}/{SOUT_REF2}", FMT_USD2, bold=True)
    ws[f"C{r}"].fill = Fill(NAVY); ws[f"C{r}"].font = F(bold=True, size=16, color="FFFFFF")
    IMPLIED_CELL = f"C{r}"
    CELLS["DCF_IMPLIED_PRICE"] = f"'DCF Valuation'!{IMPLIED_CELL}"

    r += 1; row_h(ws, r, 17)
    lbl(ws, f"B{r}", "  Current Stock Price ($)")
    cal(ws, f"C{r}", f"={A['A_PRICE']}", FMT_USD2)
    CUR_PRICE_REF = f"C{r}"

    r += 1; row_h(ws, r, 17)
    lbl(ws, f"B{r}", "  Upside / (Downside) %", bold=True, bg=LIGHT_BLUE)
    cal(ws, f"C{r}",
        f"=({IMPLIED_CELL}-{CUR_PRICE_REF})/{CUR_PRICE_REF}", FMT_PCT2, bold=True)

    # Composition
    r += 1; dsec(r, "VALUATION COMPOSITION")
    r += 1; row_h(ws, r, 17)
    lbl(ws, f"B{r}", "  PV of FCFs as % of Enterprise Value")
    cal(ws, f"C{r}", f"=IF({EV_CELL}=0,0,{SUM_PV_REF}/{EV_CELL})", FMT_PCT)
    r += 1; row_h(ws, r, 17)
    lbl(ws, f"B{r}", "  PV of Terminal Value as % of Enterprise Value")
    cal(ws, f"C{r}", f"=IF({EV_CELL}=0,0,{PV_TV_REF}/{EV_CELL})", FMT_PCT)
    r += 1; row_h(ws, r, 17)
    lbl(ws, f"B{r}", "  Implied EV / EBITDA Multiple (LTM)")
    IS_EBITDA_LTM = f"'Income Statement'!E{A['IS_EBITDA_ROW']}"
    cal(ws, f"C{r}", f"=IF({IS_EBITDA_LTM}=0,0,{EV_CELL}/{IS_EBITDA_LTM})", FMT_MULT)


# ═══════════════════════════════════════════════════════════════════════════════
# SENSITIVITY SHEET
# ═══════════════════════════════════════════════════════════════════════════════
def build_sensitivity(ws, CELLS):
    A = CELLS
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 16

    ws.merge_cells("B1:L1")
    sc(ws, "B1", "SENSITIVITY ANALYSIS  -  Implied Share Price vs Key Drivers",
       bold=True, size=13, fc="FFFFFF", bg=NAVY, al=C)
    row_h(ws, 1, 26)
    ws.merge_cells("B2:L2")
    sc(ws, "B2",
       "  Green = above current stock price (upside)  |  Red = below (downside)  |"
       "  Base case uses WACC and TGR from WACC and Assumptions sheets.",
       size=9, italic=True, fc="404040", bg=LIGHT_BLUE, al=L)
    row_h(ws, 2, 13)

    IMPLIED = A["DCF_IMPLIED_PRICE"]  # e.g. "'DCF Valuation'!C22"
    WACC_C  = A["DCF_WACC_REF"]      # e.g. "'DCF Valuation'!C4"
    TVG_C   = A["DCF_TVG_REF"]       # e.g. "'DCF Valuation'!C5"
    CUR_PX  = A["A_PRICE"]           # e.g. "Assumptions!C7"

    IS_FCFF_ROW  = A["IS_FCFF_ROW"]
    IS_EBITDA_ROW = A["IS_EBITDA_ROW"]
    IS_EBITDA_Y10 = A["IS_EBITDA_Y10"]
    SOUT = A["A_SHARES"]
    NETD = f"({A['A_DEBT']}-{A['A_CASH']})"

    WACC_OFFSETS = [-0.0150, -0.0100, -0.0050, 0.0000, +0.0050, +0.0100, +0.0150]
    TVG_OFFSETS  = [-0.0100, -0.0050, 0.0000, +0.0050, +0.0100, +0.0150]

    for ci in range(3, 10):
        ws.column_dimensions[get_column_letter(ci)].width = 13

    r = 4
    ws.merge_cells(f"B{r}:I{r}")
    sc(ws, f"B{r}",
       "  TABLE 1 - Implied Share Price  vs  WACC  x  Terminal Growth Rate",
       bold=True, size=10, fc="FFFFFF", bg=DARK_BLUE, al=L, bdr=thin_border)
    row_h(ws, r, 16)

    # Header: TGR across columns
    r += 1; row_h(ws, r, 16)
    sc(ws, f"B{r}", "  WACC \\ TGR", bold=True, size=8, fc="FFFFFF",
       bg=MED_BLUE, al=C, bdr=thin_border)
    TGR_HDR_ROW = r
    for j, offset in enumerate(TVG_OFFSETS):
        c = get_column_letter(3 + j)
        ws.column_dimensions[c].width = 13
        cal(ws, f"{c}{r}", f"={TVG_C}+({offset})", FMT_PCT2)
        ws[f"{c}{r}"].fill = Fill(MED_BLUE)
        ws[f"{c}{r}"].font = F(bold=True, size=9, color="FFFFFF")

    DATA1_START = r + 1
    for i, w_off in enumerate(WACC_OFFSETS):
        r += 1; row_h(ws, r, 17)
        # WACC label
        cal(ws, f"B{r}", f"={WACC_C}+({w_off})", FMT_PCT2)
        ws[f"B{r}"].fill = Fill(DARK_BLUE)
        ws[f"B{r}"].font = F(bold=True, size=9, color="FFFFFF")
        for j, t_off in enumerate(TVG_OFFSETS):
            c = get_column_letter(3 + j)
            wacc_c = f"B{r}"
            tgr_c  = f"{get_column_letter(3+j)}{TGR_HDR_ROW}"
            # Full DCF formula parameterized by WACC and TGR
            # SUMPRODUCT discounts each year's FCF, then adds GGM terminal value
            fcf_range = f"'Income Statement'!F{IS_FCFF_ROW}:O{IS_FCFF_ROW}"
            pv_fcf = (
                f"SUMPRODUCT({fcf_range},"
                f"1/(1+{wacc_c})^{{1,2,3,4,5,6,7,8,9,10}})"
            )
            # Y10 FCF (col O = 10th projected year)
            fcf_y10 = f"'Income Statement'!O{IS_FCFF_ROW}"
            tv_ggm  = f"{fcf_y10}*(1+{tgr_c})/MAX({wacc_c}-{tgr_c},0.0001)"
            pv_tv   = f"{tv_ggm}/(1+{wacc_c})^10"
            formula = (
                f"=IFERROR(({pv_fcf}+{pv_tv}-{NETD})/{SOUT},\"—\")"
            )
            cal(ws, f"{c}{r}", formula, FMT_USD2)
    DATA1_END = r

    # Color scale
    from openpyxl.formatting.rule import ColorScaleRule
    cs1 = ColorScaleRule(
        start_type="min",  start_color="FFC7CE",
        mid_type="percentile", mid_value=50, mid_color="FFEB9C",
        end_type="max",    end_color="C6EFCE"
    )
    rng1 = f"C{DATA1_START}:{get_column_letter(2+len(TVG_OFFSETS))}{DATA1_END}"
    ws.conditional_formatting.add(rng1, cs1)

    # Table 2: EBITDA margin vs Revenue growth
    r += 2
    ws.merge_cells(f"B{r}:H{r}")
    sc(ws, f"B{r}",
       "  TABLE 2 - Implied Share Price  vs  EBITDA Margin  x  Y1 Revenue Growth",
       bold=True, size=10, fc="FFFFFF", bg=DARK_BLUE, al=L, bdr=thin_border)
    row_h(ws, r, 16)

    MARGIN_OFFSETS = [-0.04, -0.02, 0.00, +0.02, +0.04]
    GROWTH_OFFSETS = [-0.03, -0.02, -0.01, 0.00, +0.01, +0.02, +0.03]

    r += 1; row_h(ws, r, 16)
    sc(ws, f"B{r}", "  Rev Gr\\EBITDA%", bold=True, size=8, fc="FFFFFF",
       bg=MED_BLUE, al=C, bdr=thin_border)
    MRG_HDR_ROW = r
    A_EBITDA_M_REF = A["A_EBITDA_M"]
    A_G1_REF       = A["A_G1"]
    for j, offset in enumerate(MARGIN_OFFSETS):
        c = get_column_letter(3 + j)
        ws.column_dimensions[c].width = 13
        cal(ws, f"{c}{r}", f"={A_EBITDA_M_REF}+({offset})", FMT_PCT)
        ws[f"{c}{r}"].fill = Fill(MED_BLUE)
        ws[f"{c}{r}"].font = F(bold=True, size=9, color="FFFFFF")

    DATA2_START = r + 1
    for i, g_off in enumerate(GROWTH_OFFSETS):
        r += 1; row_h(ws, r, 17)
        cal(ws, f"B{r}", f"={A_G1_REF}+({g_off})", FMT_PCT)
        ws[f"B{r}"].fill = Fill(DARK_BLUE)
        ws[f"B{r}"].font = F(bold=True, size=9, color="FFFFFF")
        for j, m_off in enumerate(MARGIN_OFFSETS):
            c       = get_column_letter(3 + j)
            g_cell  = f"B{r}"
            m_cell  = f"{get_column_letter(3+j)}{MRG_HDR_ROW}"
            base_implied = IMPLIED
            base_g       = A_G1_REF
            base_m       = A_EBITDA_M_REF
            # Ratio adjustment: adjust implied price proportionally to revenue and margin changes
            formula = (
                f"=IFERROR({base_implied}"
                f"*((1+{g_cell})/(1+{base_g}))"
                f"*({m_cell}/{base_m})"
                f",\"—\")"
            )
            cal(ws, f"{c}{r}", formula, FMT_USD2)
    DATA2_END = r

    ws.conditional_formatting.add(
        f"C{DATA2_START}:{get_column_letter(2+len(MARGIN_OFFSETS))}{DATA2_END}",
        cs1
    )

    # Football field
    r += 2
    ws.merge_cells(f"B{r}:H{r}")
    sc(ws, f"B{r}", "  FOOTBALL FIELD  -  Valuation Range Summary",
       bold=True, size=10, fc="FFFFFF", bg=DARK_BLUE, al=L, bdr=thin_border)
    row_h(ws, r, 16)

    r += 1; row_h(ws, r, 15)
    for c, hdr_t in [("B","Method"), ("C","Low ($)"), ("D","High ($)"),
                      ("E","Current ($)"), ("F","Notes")]:
        ws.column_dimensions[c].width = 18 if c in ["B","F"] else 13
        hdr_cell(ws, f"{c}{r}", hdr_t, bg=MED_BLUE, size=9)

    football_rows = [
        ("DCF  (base case)",         IMPLIED,            IMPLIED,                     CUR_PX, "Gordon Growth TV"),
        ("DCF  (bull: FCF+15%)",     f"{IMPLIED}*1.12",  f"{IMPLIED}*1.22",           CUR_PX, "Upside scenario"),
        ("DCF  (bear: FCF-15%)",     f"{IMPLIED}*0.78",  f"{IMPLIED}*0.90",           CUR_PX, "Downside scenario"),
        ("EV/EBITDA  (18-22x LTM)",  f"={IS_EBITDA_Y10}*18/{SOUT}", f"={IS_EBITDA_Y10}*22/{SOUT}", CUR_PX, "Exit multiple"),
        ("Current price  (mkt)",     CUR_PX,             CUR_PX,                      CUR_PX, "Market quote"),
    ]
    for method, low, high, cur, note in football_rows:
        r += 1; row_h(ws, r, 17)
        lbl(ws, f"B{r}", f"  {method}")
        low_formula  = f"={low}"  if not str(low).startswith("=") else low
        high_formula = f"={high}" if not str(high).startswith("=") else high
        cal(ws, f"C{r}", low_formula,  FMT_USD2)
        cal(ws, f"D{r}", high_formula, FMT_USD2)
        cal(ws, f"E{r}", f"={cur}", FMT_USD2)
        lbl(ws, f"F{r}", f"  {note}")


# ─── ENTRY POINT ──────────────────────────────────────────────────────────────
if __name__ == "__main__":
    ticker = sys.argv[1].upper() if len(sys.argv) > 1 else None

    if ticker is None:
        ans = input("  Enter ticker to auto-populate (or press Enter to skip): ").strip().upper()
        ticker = ans or None

    stamp = datetime.datetime.now().strftime("%Y-%m-%d_%I-%M%p")

    if ticker:
        name_options = [
            f"DCF_Model_{ticker}_{stamp}.xlsx",
            f"DCF_Model_{ticker}.xlsx",
            f"DCF_Model_{stamp}.xlsx",
            "DCF_Model.xlsx",
        ]
    else:
        name_options = [
            f"DCF_Model_{stamp}.xlsx",
            "DCF_Model.xlsx",
        ]

    print("\n  Save as:")
    for i, name in enumerate(name_options, 1):
        print(f"    [{i}] {name}")
    raw = input(f"  Choice [1]: ").strip()
    choice = int(raw) if raw.isdigit() and 1 <= int(raw) <= len(name_options) else 1
    filename = name_options[choice - 1]

    path = build_dcf(filename=filename)
    print(f"\n  DCF Model created: {path}")

    if ticker:
        from update_dcf import fetch_stock_data, update_excel
        print(f"  Fetching data for {ticker} from Yahoo Finance...")
        data = fetch_stock_data(ticker)
        update_excel(path, data)

    print(f"\n  Next steps:")
    print(f"    1. Open {filename} in Excel")
    print( "    2. Edit gold cells in 'Assumptions' tab")
    print( "    3. View implied price on 'DCF Valuation' tab")
    print( "    4. Explore scenarios in 'Sensitivity' tab")
